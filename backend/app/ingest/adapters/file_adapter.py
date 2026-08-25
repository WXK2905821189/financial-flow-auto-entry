"""文件导入采集适配器（决策 D6 兜底）。

读取网银导出的 CSV / Excel 报表，映射到统一流水契约；XML 预留映射点。
字段映射以「预定义标准字段 + 按实际对齐」（决策 D4=B），列名别名可配置。
"""
from __future__ import annotations

import csv
import json
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.ingest.adapters.base import BaseAdapter
from app.ingest.adapters.registry import register_adapter
from app.core.contract import BankTransaction, Direction, SourceType

# 网银报表列别名 → 契约字段 的默认映射（可按银行在 column_map 覆盖扩展）
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "txn_no": ("流水号", "交易流水号", "银行流水号", "回单号", "txn_no", "serial_no", "trans_no"),
    "txn_date": ("交易日期", "交易日", "记账日期", "入账日期", "txn_date", "date"),
    "txn_time": ("交易时间", "发生时间", "txn_time", "time"),
    "amount": ("金额", "交易金额", "发生额", "交易额", "amount", "amt"),
    "debit_amount": ("借方金额", "借方发生额", "借方", "支出金额", "debit_amount"),
    "credit_amount": ("贷方金额", "贷方发生额", "贷方", "收入金额", "credit_amount"),
    "direction": ("借贷方向", "借贷标志", "收支方向", "收付标志", "收/支", "dc_flag", "direction", "flag"),
    "counterparty_name": ("对方户名", "对方账户名", "对方名称", "对方单位名称", "收(付)方名称", "counterparty_name", "payer"),
    "counterparty_account": ("对方账号", "对方账户", "对方帐号", "收(付)方账号", "counterparty_account"),
    "summary": ("摘要", "银行摘要", "用途", "附言", "备注", "交易摘要", "summary", "remark", "memo"),
    "account_no": ("本方账号", "本方账户", "账户", "账号", "account", "account_no"),
    "currency": ("币种", "币别", "货币", "currency"),
}

DC_SYNONYMS: dict[str, Direction] = {
    "D": Direction.DEBIT,
    "C": Direction.CREDIT,
    "借": Direction.DEBIT,
    "贷": Direction.CREDIT,
    "借方": Direction.DEBIT,
    "贷方": Direction.CREDIT,
    "支出": Direction.DEBIT,
    "收入": Direction.CREDIT,
    "支": Direction.DEBIT,
    "收": Direction.CREDIT,
    "1": Direction.DEBIT,
    "2": Direction.CREDIT,
    "debit": Direction.DEBIT,
    "credit": Direction.CREDIT,
}

# 币种中文名 → ISO4217（网银原始报文常见的币种字符串）
CURRENCY_SYNONYMS: dict[str, str] = {
    "人民币": "CNY", "RMB": "CNY",
    "美元": "USD", "美金": "USD",
    "港币": "HKD",
    "欧元": "EUR",
    "英镑": "GBP",
    "日元": "JPY",
}


class _SkipReason:
    NO_TXN_NO = "缺流水号"
    AMOUNT = "金额解析失败"
    DIRECTION = "借贷方向缺失"
    CONFLICT = "借贷金额并存冲突"
    DATE = "交易日期缺失/解析失败"
    JSON_DECODE = "JSON 解析失败"
    OTHER = "其他解析失败"


def _skip_reason(exc: Exception) -> str:
    msg = str(exc)
    if msg.startswith("无法解析金额"):
        return _SkipReason.AMOUNT
    if "借、贷金额同时有值" in msg:
        return _SkipReason.CONFLICT
    if "借贷方向" in msg or "方向" in msg:
        return _SkipReason.DIRECTION
    if "缺少交易日期" in msg or "无法解析日期" in msg:
        return _SkipReason.DATE
    return _SkipReason.OTHER


def _note_skip(diagnostics: dict | None, reason: str) -> None:
    """回填跳过行统计与原因归类；`diagnostics` 为 None 时不记录（向后兼容）。"""
    if diagnostics is None:
        return
    diag = diagnostics.setdefault("skipped", {"total": 0, "reasons": {}})
    diag["total"] += 1
    diag["reasons"][reason] = diag["reasons"].get(reason, 0) + 1


@register_adapter
class FileAdapter(BaseAdapter):
    source_type = SourceType.FILE

    def fetch(
        self,
        *,
        content: bytes,
        filename: str = "flow.csv",
        bank_code: str = "CITIC",
        account_no: str = "",
        column_map: dict[str, str] | None = None,
        diagnostics: dict | None = None,
        **kwargs,
    ) -> list[BankTransaction]:
        """解析网银导出文件 → 统一契约；`diagnostics` 非空时回填跳过统计（供调用方可见）。"""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"
        if ext == "json":
            return _read_json(
                content, bank_code=bank_code, account_no=account_no, diagnostics=diagnostics
            )
        rows = _read_excel(content) if ext in ("xlsx", "xlsm", "xls") else _read_csv(content)
        if not rows:
            return []

        header = rows[0]
        idx = _resolve_columns(header, column_map)
        result: list[BankTransaction] = []
        for raw in rows[1:]:
            if not raw or all(v in (None, "") for v in raw):
                continue  # 纯空行不计入可观测统计
            try:
                txn = _map_row(raw, idx, header, bank_code, account_no)
            except (ValueError, InvalidOperation) as exc:
                _note_skip(diagnostics, exc)
                continue
            if txn is None:
                _note_skip(diagnostics, _SkipReason.NO_TXN_NO)
                continue
            result.append(txn)
        return result


def _read_json(
    content: bytes,
    *,
    bank_code: str,
    account_no: str,
    diagnostics: dict | None = None,
) -> list[BankTransaction]:
    """读取 JSON 契约 / 银行原始报文 / 批次数组，统一归一到契约。

    支持三种结构：
    1) 银行原始报文：{"account_no": "...", "flows": [{中文字段...}, ...]}
    2) 批次数组：    [{"flows": [{契约字段...}, ...]}, ...]
    3) 平铺契约数组：[{txn_no, txn_date, ...}, ...]
    """
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except (ValueError, UnicodeDecodeError):
        _note_skip(diagnostics, _SkipReason.JSON_DECODE)
        return []

    def map_flows(flows, acct):
        out: list[BankTransaction] = []
        for f in flows:
            if not isinstance(f, dict):
                continue
            try:
                txn = _map_dict(f, bank_code, acct)
            except (ValueError, InvalidOperation) as exc:
                _note_skip(diagnostics, exc)
                continue
            if txn is None:
                _note_skip(diagnostics, _SkipReason.NO_TXN_NO)
                continue
            out.append(txn)
        return out

    if isinstance(data, dict) and "flows" in data:
        acct = str(data.get("account_no") or account_no or "").strip()
        return map_flows(data["flows"], acct)

    if isinstance(data, list):
        result: list[BankTransaction] = []
        for item in data:
            if isinstance(item, dict) and "flows" in item:
                acct = str(item.get("account_no") or account_no or "").strip()
                result.extend(map_flows(item["flows"], acct))
            elif isinstance(item, dict):
                try:
                    txn = _map_dict(item, bank_code, account_no)
                except (ValueError, InvalidOperation):
                    continue
                if txn is not None:
                    result.append(txn)
        return result

    return []


def _map_dict(obj: dict, bank_code: str, account_no: str) -> BankTransaction | None:
    """将 dict 行（契约字段或银行中文字段）映射为统一契约。"""
    src = {str(k).strip(): v for k, v in obj.items()}

    def get(field: str):
        for alias in COLUMN_ALIASES.get(field, ()):
            if alias in src:
                return src[alias]
        return ""

    txn_no = str(get("txn_no") or "").strip()
    if not txn_no:
        return None

    amount, dc = _resolve_amount_direction(get)
    txn_date = _parse_date(get("txn_date"))

    return BankTransaction(
        bank_code=bank_code,
        account_no=account_no or str(get("account_no") or "").strip(),
        txn_no=txn_no,
        txn_date=txn_date,
        txn_time=_parse_time(get("txn_date"), get("txn_time")),
        currency=_parse_currency(get("currency")),
        amount=amount,
        dc_flag=dc,
        counterparty_name=str(get("counterparty_name") or "").strip() or None,
        counterparty_account=str(get("counterparty_account") or "").strip() or None,
        summary=str(get("summary") or "").strip() or None,
        ext={"source_row": obj},
    )


def _parse_currency(v) -> str:
    s = str(v or "").strip()
    if not s:
        return "CNY"
    if s in CURRENCY_SYNONYMS:
        return CURRENCY_SYNONYMS[s]
    return s.upper()


def _read_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [list(r) for r in csv.reader(StringIO(text))]


def _read_excel(content: bytes) -> list[list[str]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [_cell_str(row) for row in ws.iter_rows(values_only=True)]


def _cell_str(row: tuple) -> list[str]:
    out = []
    for v in row:
        if v is None:
            out.append("")
        elif isinstance(v, datetime):
            out.append(v.isoformat())
        else:
            out.append(str(v))
    return out


def _resolve_columns(header: list[str], column_map: dict[str, str] | None) -> dict[str, int]:
    norm = {h.strip(): i for i, h in enumerate(header)}
    idx: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        if column_map and field in column_map and column_map[field] in norm:
            idx[field] = norm[column_map[field]]
            continue
        for alias in aliases:
            if alias in norm:
                idx[field] = norm[alias]
                break
    return idx


def _map_row(
    row: list[str],
    idx: dict[str, int],
    header: list[str],
    bank_code: str,
    account_no: str,
) -> BankTransaction | None:
    def get(field: str):
        i = idx.get(field)
        if i is None or i >= len(row):
            return ""
        return row[i].strip() if isinstance(row[i], str) else row[i]

    txn_no = str(get("txn_no"))
    if not txn_no:
        return None  # 无流水号，无法归一化

    amount, dc = _resolve_amount_direction(get)
    txn_date = _parse_date(get("txn_date"))

    return BankTransaction(
        bank_code=bank_code,
        account_no=account_no or str(get("account_no")),
        txn_no=txn_no,
        txn_date=txn_date,
        txn_time=_parse_time(get("txn_date"), get("txn_time")),
        currency=_parse_currency(get("currency")),
        amount=amount,
        dc_flag=dc,
        counterparty_name=str(get("counterparty_name")) or None,
        counterparty_account=str(get("counterparty_account")) or None,
        summary=str(get("summary")) or None,
        ext={"source_row": row, "header": header},
    )


def _parse_amount(v) -> Decimal:
    s = str(v or "0").replace(",", "").replace("¥", "").replace("￥", "").replace(" ", "")
    try:
        return abs(Decimal(s))
    except InvalidOperation:
        raise ValueError(f"无法解析金额：{v!r}")


def _parse_direction(v: str, amount_v: str) -> Direction:
    key = str(v or "").strip()
    if key in DC_SYNONYMS:
        return DC_SYNONYMS[key]
    # 方向列缺失：按金额负号推断
    if str(amount_v or "").strip().startswith("-"):
        return Direction.DEBIT
    raise ValueError(f"无法解析借贷方向：{v!r}")


def _resolve_amount_direction(get) -> tuple[Decimal, Direction]:
    """解析金额与方向：优先常规「金额」单列，回退到「借方/贷方金额」分列（招商网银格式）。"""
    if str(get("amount") or "").strip():
        raw = str(get("amount") or "").strip()
        return _parse_amount(get("amount")), _parse_direction(get("direction"), raw)

    debit = str(get("debit_amount") or "").strip()
    credit = str(get("credit_amount") or "").strip()
    if debit and credit:
        raise ValueError("借、贷金额同时有值，无法归一化方向")
    if debit:
        return _parse_amount(get("debit_amount")), Direction.DEBIT
    if credit:
        return _parse_amount(get("credit_amount")), Direction.CREDIT
    raise ValueError("缺少金额字段")


def _parse_date(v) -> date:
    s = str(v or "").strip()
    if not s:
        raise ValueError("缺少交易日期")
    if "T" in s or " " in s:
        s = s.replace("T", " ").split(" ")[0]
    s = s.replace("/", "-").replace(".", "-")
    try:
        return date.fromisoformat(s)
    except ValueError:
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except ValueError:
            raise ValueError(f"无法解析日期：{v!r}")


def _parse_time(date_v, time_v) -> time | None:
    s = str(time_v or "").strip()
    if s:
        try:
            return time.fromisoformat(s)
        except ValueError:
            pass
    dv = str(date_v or "")
    if " " in dv or "T" in dv:
        try:
            return datetime.fromisoformat(dv.replace("T", " ")).time()
        except ValueError:
            pass
    return None